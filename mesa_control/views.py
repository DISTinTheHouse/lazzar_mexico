from django.shortcuts import render, HttpResponse, redirect
from django.db.models import F, Q
import datetime
from django.contrib.auth.decorators import login_required
from mainapp.models import CorreosLazzar
import MySQLdb
import pyodbc
from datetime import datetime, timedelta
from django.contrib.auth.models import User
import pytz
#EXCEL
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse
# MEMORIA CACHE
# MEMORIA CACHE
from django.views.decorators.cache import cache_page
from django.core.cache import cache

# CONEXION A MYSQL DE PROSCAI
DB_HOST = 'as.galio.net' 
DB_USER = 'u129' 
DB_PASS = 'lbrfyg' 
DB_NAME = 'db129lazzar'
DB_PORT = 61129

# CONEXION SQLSERVER BITACOR DE PRODUCCION
server = 'lazzarbodegavpn.dyndns.info'
database = 'CODIGO_BARRAS'
username = 'sa'
password = 'lazzar2018'

# EJECUTA EL QUERY Y RETORNA LA INFORMACIÓN
def run_query(query=''): 
    datos = [DB_HOST, DB_USER, DB_PASS, DB_NAME, DB_PORT] 
    
    conn = MySQLdb.connect(*datos) # Conectar a la base de datos 
    cursor = conn.cursor()         # Crear un cursor 
    cursor.execute(query)          # Ejecutar una consulta 

    if query.upper().startswith('SELECT'): 
        data = cursor.fetchall()   # Traer los resultados de un select 
    else: 
        conn.commit()              # Hacer efectiva la escritura de datos 
        data = None 
    
    cursor.close()                 # Cerrar el cursor 
    conn.close()                   # Cerrar la conexión 

    return data


# EJECUTA EL QUERY Y RETORNA LA INFORMACIÓN (SQLSERVER)
def run_query_sql(query='', connection_timeout_seconds=380):
    #cadena_conexion = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
    cadena_conexion = f'DSN=mydsn;SERVER={server};DATABASE={database};UID={username};PWD={password}'
    conexion = pyodbc.connect(cadena_conexion, timeout=connection_timeout_seconds)  # Establecer un tiempo límite para la conexión
    cursor = conexion.cursor()

    cursor.execute(query)
    data = cursor.fetchall()

    conexion.commit()  # Siempre es recomendable confirmar los cambios antes de cerrar la conexión.
    conexion.close()  # Cerrar la conexión
    
    return data  


# Create your views here.
@login_required(login_url="login")
def estatus_mesa(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    
    if usuario.esvendedor:
        vendedores = CorreosLazzar.objects.filter(esvendedor=usuario.esvendedor).values_list()

        return render(request, 'restringido.html',{
        'title':'Acceso Restringido',
        'vendedores': vendedores,
        'leyenda':False,
        })
    else:
        return render(request, 'estatus_mesa.html',{
        'title':'Estatus del Pedido',
        'leyenda':False
        })

@login_required(login_url="login")
def estatus_pedido_mesa(request):
    id = request.user.id
    cache_key = f"usuario_{id}:estatus_pedido_mesa"

    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    #Verificar si la respuesta esta en el cache:
    cached_response = cache.get(cache_key)
    if cached_response:
        return cached_response

    #Realizar consultas SQL utilizando el modelo Lazzar
    # *********************PEDIDOS ACTIVOS ************
    correos = CorreosLazzar.objects.all()
    query = (f"SELECT PENUM, CLINOM, AGDESCR, PEFECHA, PEPAR3, PEFECHAEMPAQUE, PEDESDE, PEVENCE, PEPAR7, C.CLICP FROM FPLIN L "
                 + "INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG "
                 + "A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
                 + "AND P.PENUM <= 'P99999' AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;")

    results = run_query(query) 
        
    if len(results) > 0:
            result_ped = list(results)
            total = len(results)
            total_activos = len(results)
    else:
            result_ped = 0
            total = 0
            total_activos = 0


    # *********************PEDIDOS VENCIDOS ************************************************
    query = (f"SELECT PENUM FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ "
                 + "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND "
                 + "P.PENUM >= 'P' AND P.PENUM <= 'P99999' AND datediff(now(), PEDESDE) > 0  AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;")
        
    results = run_query(query) 
        
    if len(results) > 0:
            total_vencidos = len(results)
    else:
            total_vencidos = 0

    # *********************PEDIDOS VENCEN HOY ************************************************
    query = (f"SELECT PENUM FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ "
                 + "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND "
                 + "P.PENUM >= 'P' AND P.PENUM <= 'P99999' AND datediff(now(), PEDESDE) = 0  AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;")
        
    results = run_query(query) 
        
    if len(results) > 0:
            total_ven_hoy = len(results)
    else:
            total_ven_hoy = 0
    # *********************PEDIDOS VENCEN 1 SEMANA ************************************************
    query = (f"SELECT P.PENUM, P.PEPAR3, P.PEFECHA, P.PEDESDE, P.PEVENCE FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
         + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
         + "WHERE (PLCANT - PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
         + "AND P.PENUM <= 'P99999' AND DATEDIFF(NOW(), P.PEDESDE) AND DATEDIFF(NOW(), P.PEDESDE) BETWEEN -5 AND 0 AND PESPEDIDO = 1 "
         + "GROUP BY P.PENUM, P.PEPAR3 ORDER BY P.PEPAR3;")
        
    results = run_query(query) 
        
    if len(results) > 0:
            total_ven_sem = len(results)
    else:
            total_ven_sem = 0

     #PEDIDOS VENCIDOS PORCENTAJE
    resultado_porcentaje = (total_vencidos * 100 / total_activos)
    #PEDIDOS EN TIEMPO
    variante_porcentaje_dos = total_activos - total_vencidos
    resultado_porcentaje_dos = (variante_porcentaje_dos / total_activos ) * 100
    #CANTIDAD PEDIDOS EN TIEMPO
    resumen_resultado = (total_activos - total_vencidos)

    response = render(request, 'activos.html',{
        'title':'Estatus del Pedido',
        'result_ped':result_ped,
        'total_ped':total,
        'total_vencidos':total_vencidos,
        'total_ven_hoy':total_ven_hoy,
        'total_ven_sem':total_ven_sem,
        'total_activos':total_activos,
        'resultado_porcentaje': resultado_porcentaje,
        'resumen_resultado':resumen_resultado,
        'resultado_porcentaje_dos':resultado_porcentaje_dos,
        'correos':correos,
        'leyenda':True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

    #Almacenar en caché la respuesta
    cache.set(cache_key, response, 60 * 15)  
    return response

@login_required(login_url="login")
def activos(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    return render(request, 'activos.html',{
        'title':'Pedidos Activos',
        'total_ped':"",
        'total_vencidos':"",
        'total_ven_hoy':"",
        'total_ven_sem':"",
        'total_activos':"",
        'resultado_porcentaje':"",
        'resumen_resultado':"",
        'resultado_porcentaje_dos':"",
        'leyenda':False,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })
    

@login_required(login_url="login") 
def detalle(request):
        ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
        acceso_total = request.user.groups.filter(name='acceso_total').exists()
        mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
        if request.method == 'POST':
             folio = request.POST['variable']
        else:
             folio = 'detboton'
    

        query = (f"SELECT P.PENUM, P.PEFECHA, P.PEDESDE, P.PEVENCE," 
            +   "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_1,"
            +   "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_2,"
            +   "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_3,"
            +   "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_4 "
            +   "FROM FPENC P "
            +   "WHERE P.PEOTROTXT = '" + folio 
            +   "' AND P.PEPAR0 >= '6' "
            +   " AND P.PEPAR0 <= '6ZZZZZ' "
            +   "AND P.PENUM >= 'E' "
            +   "AND P.PENUM <= 'E999999999';")


        results = run_query(query)
        resultado = list(results)

        query = (f"SELECT P.PENUM, P.PEFECHA, P.PEDESDE, P.PEVENCE," 
            +   "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_1,"
            +   "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_2,"
            +   "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_3,"
            +   "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_4 "
            +   "FROM FPENC P "
            +   "WHERE P.PEOTROTXT = '" + folio 
            +   "' AND P.PEPAR0 >= '8' "
            +   " AND P.PEPAR0 <= '8ZZZZZ' "
            +   "AND P.PENUM >= 'E' "
            +   "AND P.PENUM <= 'E999999999';")

        results = run_query(query)
        segresultado = list(results)

        return render(request, 'detalle.html', {
        'title':folio,
        'resultado': resultado,
        'segresultado': segresultado,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    }) 

@login_required(login_url="login")
def vencidos(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    clasA = 0
    clasB = 0
    clasC = 0
    clasD = 0
    clasE = 0
    clasF = 0
    clasX = 0
    clas = 0
    
    query = (f"SELECT P.PENUM, P.PEPAR3, P.PEFECHA, P.PEDESDE, P.PEVENCE FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
             + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
             + "WHERE (PLCANT - PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
             + "AND P.PENUM <= 'P99999' AND DATEDIFF(NOW(), P.PEDESDE) > 0 AND PESPEDIDO = 1 GROUP BY P.PENUM, "
             + "P.PEPAR3 ORDER BY P.PEPAR3;")
        
    results = run_query(query)
    result_ped = list(results)

    for resultado in result_ped:
        if resultado[1] == "3A":
            clasA += 1
        elif resultado[1] == "3B":
            clasB += 1
        elif resultado[1] == "3C":
            clasC += 1
        elif resultado[1] == "3D":
            clasD += 1
        elif resultado[1] == "3E":
            clasE += 1
        elif resultado[1] == "3F":
            clasF += 1
        elif resultado[1] == "3X":
            clasX += 1
        else:
            clas += 1

    return render(request, 'vencidos.html',{
        'title':'Pedidos Vencidos',
        'resultado': result_ped,
        'clasA':clasA,
        'clasB':clasB,
        'clasC':clasC,
        'clasD':clasD,
        'clasE':clasE,
        'clasF':clasF,
        'clasX':clasX,
        'clas':clas,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="dashboard")
def dashboard (request):
     ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
     acceso_total = request.user.groups.filter(name='acceso_total').exists()
     mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
     clasA = 0
     clasB = 0
     clasC = 0
     clasD = 0
     clasE = 0
     clasF = 0
     clasX = 0
     clas = 0

     query = (f"SELECT P.PENUM, P.PEPAR3, P.PEDESDE, P.PEVENCE FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
             + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
             + "WHERE (PLCANT - PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
             + "AND P.PENUM <= 'P99999' AND DATEDIFF(NOW(), P.PEDESDE) > 0 AND PESPEDIDO = 1 GROUP BY P.PENUM, "
             + "P.PEPAR3 ORDER BY P.PEPAR3;")
     results = run_query(query)
     results_ped = list (results)

     for resultado in results_ped:
        if resultado[1] == "3A":
            clasA += 1
        elif resultado[1] == "3B":
            clasB += 1
        elif resultado[1] == "3C":
            clasC += 1
        elif resultado[1] == "3D":
            clasD += 1
        elif resultado[1] == "3E":
            clasE += 1
        elif resultado[1] == "3F":
            clasF += 1
        elif resultado[1] == "3X":
            clasX += 1
        else:
            clas += 1

     return render(request, 'dashboard.html',{
        'title':'Pedidos Vencidos',
        'resultado': results_ped,
        'clasA':clasA,
        'clasB':clasB,
        'clasC':clasC,
        'clasD':clasD,
        'clasE':clasE,
        'clasF':clasF,
        'clasX':clasX,
        'clas':clas,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="dashboard")
def bit_produccion (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        folio = request.POST['variable']
    else:
        folio = 'detboton'
    #---------FECHA INICIO-----------
    query = (f"SELECT P.PEFECHA, P.PEFECHAEMPAQUE, P.PEDESDE, P.PEVENCE FROM FPENC P " 
            + "WHERE P.PENUM = '" + folio + "' ")
    results = run_query(query)
    fecha = list(results)

    #------CODIGOS PENDIENTES DE SURIR-------
    query = (f"SELECT MID(I.ICOD,3,10) AS CODIGO, I.IDESCR AS DESCRIPCION, (L.PLCANT-PLSURT-PLASIGNADO) AS PENDIENTE FROM FPENC P "
            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE P.PENUM = '"
            + folio + "' AND (L.PLCANT-PLSURT-PLASIGNADO) != 0 AND I.ICOD >= '1';")
    
    results = run_query(query)
    codigos = list (results)

    existencias = []
    total_pendiente = 0

    for codigo in codigos:
        query_exis = (f"SELECT existencia FROM Exitencia WHERE CODIGO = '{codigo[0]}'")
        exis_cod = run_query_sql(query_exis)

        if len(exis_cod) > 0:
            for exis in exis_cod:
                exis_val = exis[0]
        else:  
            exis_val = 0



        query = (f"SELECT PL.PLCANT-PL.PLSURT AS PENDIENTE, P.PENUM, MIN(P.PEVENCE) FROM FPLIN PL "
                + f"INNER JOIN FINV I ON I.ISEQ = PL.ISEQ INNER JOIN FFAM F ON F.FAMTNUM = I.IFAM6 "
                + f"INNER JOIN FPENC P ON P.PESEQ = PL.PESEQ WHERE MID(I.ICOD,3,10) = '{codigo[0]}' "
                + f"AND (PL.PLCANT-PL.PLSURT) > 0 AND PLTIPMV = 'E' AND P.PEINICIAL = 0 AND P.PEPAR1 = '1PW';")
        
        total_pendiente += codigo[2]

        results = run_query(query)
        producciones = list(results)

        if producciones[0][0] is None:
            query = (f"SELECT PL.PLCANT-PL.PLSURT AS PENDIENTE, P.PENUM, MIN(P.PEVENCE) FROM FPLIN PL "
                    + f"INNER JOIN FINV I ON I.ISEQ = PL.ISEQ INNER JOIN FFAM F ON F.FAMTNUM = I.IFAM6 "
                    + f"INNER JOIN FPENC P ON P.PESEQ = PL.PESEQ WHERE MID(I.ICOD,3,10) = '{codigo[0]}' "
                    + f" AND (PL.PLCANT-PL.PLSURT) > 0 AND PLTIPMV = 'O' AND P.PEINICIAL = 0 AND P.PEPAR1 = '1PW';")

            results = run_query(query)
            compras = list(results)

            existencia_data = {
                'codigo': codigo[0],
                'descripcion': codigo[1],
                'pendiente': codigo[2],
                'existencia': exis_val,
                'pendiente_produccion': compras[0][0],
                'oren': compras[0][1],
                'fecha': compras[0][2]
            }
            
        else:
            existencia_data = {
                'codigo': codigo[0],
                'descripcion': codigo[1],
                'pendiente': codigo[2],
                'existencia': exis_val,
                'pendiente_produccion': producciones[0][0],
                'oren': producciones[0][1],
                'fecha': producciones[0][2]
            }

        # Agregar los datos de existencia a la lista existencias
        existencias.append(existencia_data)

    # ---------ALMACEN-----------------
    query = (f"SELECT P.PENUM, (SELECT C.COML5 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS ALMACEN "
            + "FROM FPENC P WHERE P.PENUM = '" + folio + "';")
    
    res_comentario = run_query(query)
    almacen = list (res_comentario)
    
    total_alm = 0
    total_surt = 0
    total_pendi = 0

    query_alm = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "P.PENUM = '" + folio + "' ORDER BY I.ICOD;")

    alm = run_query(query_alm)
    detalle_alm = list(alm)

    for item_alm in detalle_alm:
        total_alm += item_alm[3]

    total_surt = sum(item[4] for item in detalle_alm)
    total_pendi = sum(item[5] for item in detalle_alm)

    # ---------produccion-----------------
    query = (f"SELECT LTRIM(RTRIM(orden)), pedidos_incluye, CONVERT(varchar,fecha_op,3) as [DD/MM/YY] , CONVERT(varchar,fecha_diseno,3) as [DD/MM/YY], "
            + "CONVERT(varchar,fecha_trazo,3) as [DD/MM/YY], CONVERT(varchar,fecha_avios,3) as [DD/MM/YY], CONVERT(varchar,fecha_tela,3) as [DD/MM/YY], "
            + "CONVERT(varchar,fecha_corte_real,3) as [DD/MM/YY], CONVERT(varchar,fecha_comp_cc,3) as [DD/MM/YY], habilitaciones, comen_prod FROM Bit_corte WHERE eliminado IS NULL AND "
            + "(PEDIDO = '"+ folio + "' OR pedidos_incluye LIKE '%" + folio + "%');")
    
    results = run_query_sql(query)
    resultado = list (results)

    largo = len(resultado)
    i = 1
    consulta = ""

    if largo > 0:
        for orden in resultado:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "' OR "
                i += 1

        detalle_prod = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")

        ops = run_query(detalle_prod)
        detalle_ops = list(ops)
       
        total_query = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") GROUP BY P.PENUM;")

        total_ops = run_query(total_query)
        totales_ops = list(total_ops)

        fechas_prod = (f"SELECT P.PENUM, P.PEVENCE FROM FPENC P INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ "
                    + "INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE (" + consulta + ") GROUP BY P.PENUM ORDER BY P.PENUM;")

        fecha_ops = run_query(fechas_prod)
        fechas_list = list(fecha_ops)
    else:
        detalle_ops = None
        totales_ops = None
        fechas_list = None

    # ---------bordados-----------------
    query = (f"SELECT P.PENUM, P.PEFECHA, P.PEDESDE, P.PEVENCE," 
            +   "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_1,"
            +   "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_2,"
            +   "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_3,"
            +   "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_4 "
            +   "FROM FPENC P "
            +   "WHERE P.PEOTROTXT = '" + folio 
            +   "' AND P.PEPAR0 >= '7' "
            +   " AND P.PEPAR0 <= '8ZZZZZ' "
            +   "AND P.PENUM >= 'E' "
            +   "AND P.PENUM <= 'E999999999';")

    results = run_query(query)
    segresultado = list(results)

    largo = len(segresultado)
    i = 1
    consulta = ""

    if largo > 0:
        for orden in segresultado:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "' OR "
                i += 1

        detalle_bord = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")
    
        ops_bord = run_query(detalle_bord)
        detalle_bordados = list(ops_bord)

        total_query_bor = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") GROUP BY P.PENUM;")
        
        total_bor = run_query(total_query_bor)
        totales_bor = list(total_bor)
    else:
        detalle_bordados = None
        totales_bor = None

    # ---------compras-----------------
    query = (f"SELECT P.PENUM, P.PEFECHA, P.PEVENCE, (SELECT C.COML5 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMPRAS "
            + "FROM FPENC P WHERE P.PENUMELLOS = '" + folio + "';")

    results = run_query(query)
    orden_com = list(results)

    largo = len(orden_com)
    i = 1
    consulta = ""

    if largo > 0:
        for oc in orden_com:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + oc[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + oc[0].strip() + "' OR "
                i += 1

        detalle_comp = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")
    
        ops_com = run_query(detalle_comp)
        detalle_compra = list(ops_com)

        total_query_comp = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                            + "(" + consulta + ") GROUP BY P.PENUM;")
        
        total_comp = run_query(total_query_comp)
        totales_comp = list(total_comp)
    else:
        detalle_compra = None
        totales_comp = None

    # ---------facturación-----------------
    query = (f"SELECT DNUM, DFECHA, (SELECT SUM(AICANTF) FROM FAXINV A INNER JOIN FINV I ON I.ISEQ = A.ISEQ "
            + "WHERE A.DSEQ = D.DSEQ AND ICOD >= '1' AND ICOD <= 'Z99999'), (SELECT SUM(AICANTF) FROM FAXINV A "
            + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE A.DSEQ = D.DSEQ AND (ICOD <= '1' OR ICOD = 'ZANTICIPO')), "
            + "DTALON,(SELECT COML4 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM), DCANCELADA FROM FDOC D WHERE DREFER = '" + folio + "' AND D.DNUM >= 'D' AND D.DNUM <= 'FT9999' ")

    results = run_query(query)
    facturas = list(results)

    largo = len(facturas)
    i = 1
    consulta = ""

    if largo > 0:
        for fact in facturas:
            if i == largo:
                consulta = consulta + "D.DNUM = '" + fact[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "D.DNUM = '" + fact[0].strip() + "' OR "
                i += 1

        detalle_fac = (f"SELECT D.DNUM, ICOD, IDESCR, AICANTF FROM FDOC D INNER JOIN FAXINV A ON A.DSEQ = D.DSEQ "
                    + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE (" + consulta + ") ORDER BY D.DNUM") 

        res_fact = run_query(detalle_fac)
        detalle_facturas = list(res_fact)

        total_query_fac = (f"SELECT D.DNUM, SUM(AICANTF) FROM FDOC D INNER JOIN FAXINV A ON A.DSEQ = D.DSEQ "
                          + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE(" + consulta + ") GROUP BY D.DNUM")
        
        total_fac = run_query(total_query_fac)
        totales_fac = list(total_fac)
    else:
        detalle_facturas = None
        totales_fac = None

    return render(request, 'detalle.html', {
        'title':folio,
        'existencias': existencias,
        'total_pendiente': total_pendiente,
        'almacen': almacen,
        'detalle_alm': detalle_alm,
        'resultado': resultado,
        'detalle_ops': detalle_ops,
        'segresultado': segresultado,
        'detalle_bordados': detalle_bordados,
        'orden_com': orden_com,
        'detalle_compra': detalle_compra,
        'facturas': facturas,
        'detalle_facturas': detalle_facturas,
        'total_alm': total_alm,
        'total_surt': total_surt,
        'total_pendi': total_pendi,
        'totales_ops': totales_ops,
        'totales_bor': totales_bor,
        'totales_comp': totales_comp,
        'totales_fac': totales_fac,
        'fecha': fecha,
        'fechas_list':fechas_list,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def faltantes (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        folio = request.POST['variable']
    else:
        folio = 'detboton'

    # ---------Pedido-----------------
    query = (f"SELECT MID(I.ICOD,3,10) AS CODIGO, I.IDESCR AS DESCRIPCION, (L.PLCANT-PLSURT-PLASIGNADO) AS PENDIENTE FROM FPENC P "
            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE P.PENUM = '"
            + folio + "' AND (L.PLCANT-PLSURT-PLASIGNADO) != 0 AND I.ICOD >= '1';")
    
    results = run_query(query)
    codigos = list (results)

    existencias = []
    total_pendiente = 0

    for codigo in codigos:
        query_exis = (f"SELECT existencia FROM Exitencia WHERE CODIGO = '{codigo[0]}'")
        exis_cod = run_query_sql(query_exis)

        if len(exis_cod) > 0:
            for exis in exis_cod:
                exis_val = exis[0]
        else:  
            exis_val = 0

        query = (f"SELECT PL.PLCANT-PL.PLSURT AS PENDIENTE, P.PENUM, MIN(P.PEVENCE) FROM FPLIN PL "
                + f"INNER JOIN FINV I ON I.ISEQ = PL.ISEQ INNER JOIN FFAM F ON F.FAMTNUM = I.IFAM6 "
                + f"INNER JOIN FPENC P ON P.PESEQ = PL.PESEQ WHERE MID(I.ICOD,3,10) = '{codigo[0]}' "
                + f"AND (PL.PLCANT-PL.PLSURT) > 0 AND PLTIPMV = 'E' AND P.PEINICIAL = 0 AND P.PEPAR1 = '1PW';")
        
        total_pendiente += codigo[2]

        results = run_query(query)
        producciones = list(results)

        if producciones[0][0] is None:
            query = (f"SELECT PL.PLCANT-PL.PLSURT AS PENDIENTE, P.PENUM, MIN(P.PEVENCE) FROM FPLIN PL "
                    + f"INNER JOIN FINV I ON I.ISEQ = PL.ISEQ INNER JOIN FFAM F ON F.FAMTNUM = I.IFAM6 "
                    + f"INNER JOIN FPENC P ON P.PESEQ = PL.PESEQ WHERE MID(I.ICOD,3,10) = '{codigo[0]}' "
                    + f" AND (PL.PLCANT-PL.PLSURT) > 0 AND PLTIPMV = 'O' AND P.PEINICIAL = 0 AND P.PEPAR1 = '1PW';")

            results = run_query(query)
            compras = list(results)

            existencia_data = {
                'codigo': codigo[0],
                'descripcion': codigo[1],
                'pendiente': codigo[2],
                'existencia': exis_val,
                'pendiente_produccion': compras[0][0],
                'oren': compras[0][1],
                'fecha': compras[0][2]
            }
            
        else:
            existencia_data = {
                'codigo': codigo[0],
                'descripcion': codigo[1],
                'pendiente': codigo[2],
                'existencia': exis_val,
                'pendiente_produccion': producciones[0][0],
                'oren': producciones[0][1],
                'fecha': producciones[0][2]
            }

        # Agregar los datos de existencia a la lista existencias
        existencias.append(existencia_data)

         # ---------produccion-----------------
    query = (f"SELECT LTRIM(RTRIM(orden)), pedidos_incluye, CONVERT(varchar,fecha_op,3) as [DD/MM/YY] , CONVERT(varchar,fecha_diseno,3) as [DD/MM/YY], "
            + "CONVERT(varchar,fecha_trazo,3) as [DD/MM/YY], CONVERT(varchar,fecha_avios,3) as [DD/MM/YY], CONVERT(varchar,fecha_tela,3) as [DD/MM/YY], "
            + "CONVERT(varchar,fecha_corte_real,3) as [DD/MM/YY], CONVERT(varchar,fecha_comp_cc,3) as [DD/MM/YY], comen_prod FROM Bit_corte WHERE eliminado IS NULL AND "
            + "(PEDIDO = '"+ folio + "' OR pedidos_incluye LIKE '%" + folio + "%');")
    
    results = run_query_sql(query)
    resultado = list (results)

    largo = len(resultado)
    i = 1
    consulta = ""

    if largo > 0:
        for orden in resultado:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "' OR "
                i += 1

        detalle_prod = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")

        ops = run_query(detalle_prod)
        detalle_ops = list(ops)
       
        total_query = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") GROUP BY P.PENUM;")

        total_ops = run_query(total_query)
        totales_ops = list(total_ops)
    else:
        detalle_ops = None
        totales_ops = None

    # ---------bordados-----------------
    query = (f"SELECT P.PENUM, P.PEFECHA, P.PEDESDE, P.PEVENCE," 
            +   "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_1,"
            +   "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_2,"
            +   "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_3,"
            +   "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_4 "
            +   "FROM FPENC P "
            +   "WHERE P.PEOTROTXT = '" + folio 
            +   "' AND P.PEPAR0 >= '7' "
            +   " AND P.PEPAR0 <= '8ZZZZZ' "
            +   "AND P.PENUM >= 'E' "
            +   "AND P.PENUM <= 'E999999999';")

    results = run_query(query)
    segresultado = list(results)

    largo = len(segresultado)
    i = 1
    consulta = ""

    if largo > 0:
        for orden in segresultado:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "' OR "
                i += 1

        detalle_bord = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")
    
        ops_bord = run_query(detalle_bord)
        detalle_bordados = list(ops_bord)

        total_query_bor = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") GROUP BY P.PENUM;")
        
        total_bor = run_query(total_query_bor)
        totales_bor = list(total_bor)
    else:
        detalle_bordados = None
        totales_bor = None

    # ---------compras-----------------
    query = (f"SELECT P.PENUM, P.PEFECHA, P.PEVENCE, (SELECT C.COML5 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMPRAS "
            + "FROM FPENC P WHERE P.PENUMELLOS = '" + folio + "';")

    results = run_query(query)
    orden_com = list(results)

    largo = len(orden_com)
    i = 1
    consulta = ""

    if largo > 0:
        for oc in orden_com:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + oc[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + oc[0].strip() + "' OR "
                i += 1

        detalle_comp = (f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
                + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                + "(" + consulta + ") ORDER BY P.PENUM;")
    
        ops_com = run_query(detalle_comp)
        detalle_compra = list(ops_com)

        total_query_comp = (f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
                            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
                            + "(" + consulta + ") GROUP BY P.PENUM;")
        
        total_comp = run_query(total_query_comp)
        totales_comp = list(total_comp)
    else:
        detalle_compra = None
        totales_comp = None

    # ---------facturación-----------------
    query = (f"SELECT DNUM, DFECHA, (SELECT SUM(AICANTF) FROM FAXINV A INNER JOIN FINV I ON I.ISEQ = A.ISEQ "
            + "WHERE A.DSEQ = D.DSEQ AND ICOD >= '1' AND ICOD <= 'Z99999'), (SELECT SUM(AICANTF) FROM FAXINV A "
            + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE A.DSEQ = D.DSEQ AND (ICOD <= '1' OR ICOD = 'ZANTICIPO')), "
            + "DTALON, DCANCELADA FROM FDOC D WHERE DREFER = '" + folio + "' AND D.DNUM >= 'D' AND D.DNUM <= 'FT9999' ;")

    results = run_query(query)
    facturas = list(results)

    largo = len(facturas)
    i = 1
    consulta = ""

    if largo > 0:
        for fact in facturas:
            if i == largo:
                consulta = consulta + "D.DNUM = '" + fact[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "D.DNUM = '" + fact[0].strip() + "' OR "
                i += 1

        detalle_fac = (f"SELECT D.DNUM, ICOD, IDESCR, AICANTF FROM FDOC D INNER JOIN FAXINV A ON A.DSEQ = D.DSEQ "
                    + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE (" + consulta + ") ORDER BY D.DNUM") 

        res_fact = run_query(detalle_fac)
        detalle_facturas = list(res_fact)

        total_query_fac = (f"SELECT D.DNUM, SUM(AICANTF) FROM FDOC D INNER JOIN FAXINV A ON A.DSEQ = D.DSEQ "
                          + "INNER JOIN FINV I ON I.ISEQ = A.ISEQ WHERE(" + consulta + ") GROUP BY D.DNUM")
        
        total_fac = run_query(total_query_fac)
        totales_fac = list(total_fac)
    else:
        detalle_facturas = None
        totales_fac = None

    return render(request, 'faltantes.html', {
        'title':folio,
        'existencias': existencias,
        'total_pendiente': total_pendiente,
        'resultado': resultado,
        'detalle_ops': detalle_ops,
        'segresultado': segresultado,
        'detalle_bordados': detalle_bordados,
        'orden_com': orden_com,
        'detalle_compra': detalle_compra,
        'facturas': facturas,
        'detalle_facturas': detalle_facturas,
        'totales_comp': totales_comp,
        'totales_bor': totales_bor,
        'totales_ops': totales_ops,
        'totales_fac': totales_fac,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def vencidos_semana(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        desde_fecha = request.POST['desde_fecha']
        hasta_fecha = request.POST['hasta_fecha']
        clasA = 0
        clasB = 0
        clasC = 0
        clasD = 0
        clasE = 0
        clasF = 0
        clasX = 0
        clas = 0
        
        query = (
            f"SELECT P.PENUM, P.PEPAR3, P.PEFECHA, P.PEDESDE, P.PEVENCE "
            f"FROM FPLIN L "
            f"INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
            f"INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ "
            f"INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
            f"WHERE (PLCANT - PLSURT) <> 0 "
            f"AND PEFECHA >= '2023-01-01' "
            f"AND P.PENUM >= 'P' AND P.PENUM <= 'P99999' "
            f"AND P.PEDESDE BETWEEN '{desde_fecha}' AND '{hasta_fecha}' "
            f"AND PESPEDIDO = 1 "
            f"GROUP BY P.PENUM, P.PEPAR3 "
            f"ORDER BY P.PEPAR3;"
        )

            
        results = run_query(query)
        result_ped = list(results)

        for resultado in result_ped:
            if resultado[1] == "3A":
                clasA += 1
            elif resultado[1] == "3B":
                clasB += 1
            elif resultado[1] == "3C":
                clasC += 1
            elif resultado[1] == "3D":
                clasD += 1
            elif resultado[1] == "3E":
                clasE += 1
            elif resultado[1] == "3F":
                clasF += 1
            elif resultado[1] == "3X":
                clasX += 1
            else:
                clas += 1

        return render(request, 'vencidos_semana.html',{
            'title':'Pedidos Vencidos',
            'resultado': result_ped,
            'desde_fecha': desde_fecha,
            'hasta_fecha': hasta_fecha,
            'clasA':clasA,
            'clasB':clasB,
            'clasC':clasC,
            'clasD':clasD,
            'clasE':clasE,
            'clasF':clasF,
            'clasX':clasX,
            'clas':clas,
            'leyenda': True,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render(request, 'vencidos_semana.html', {
            'leyenda': False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required(login_url="login")
def existencias_mesa (request):
    return render (request, 'existencias_mesa.html')

@login_required(login_url="login")
def procesos (request):
    return render (request, "procesos.html")

@login_required(login_url="login")
def factura(request):
    resultado = []
    start_fecha = ""
    end_fecha = ""
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    
    if request.method == 'POST':
        start_fecha = request.POST['start_date']
        end_fecha = request.POST['end_date']

        query = (f"SELECT F.DREFER, F.DNUM, F.DBRUTO, C.CLINOM, A.AGDESCR "
                "FROM FDOC F INNER JOIN fpenc P ON PENUM = F.DREFER LEFT JOIN FPLIN M ON P.PESEQ = M.PESEQ "
                "LEFT JOIN FINV I ON I.ISEQ = M.ISEQ LEFT JOIN FCLI C ON C.CLISEQ = F.CLISEQ "
                "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 INNER JOIN FAG AA ON AA.AGTNUM = P.PEPAR4 "
                "WHERE F.DFECHA BETWEEN '" + start_fecha + "' AND '" + end_fecha + "' AND F.DITIPMV = 'F' AND F.DCANCELADA = 0 GROUP BY F.DNUM ") 

        results = run_query(query) 
        resultado = list(results)


    return render(request, 'factura.html', {
        'resultado': resultado,
        'leyenda': True,
        'start_date': start_fecha,
        'end_date': end_fecha,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def detalle_factura(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
 
    if request.method == 'POST':
        folio = request.POST['variable']  

        query = (f"SELECT F.DNUM, F.DREFER, I.IDESCR, CAST(M.PLCANT AS SIGNED) AS PLCANT, C.CLINOM, C.CLICONT, CLITEL, CLITEL4, A.AGDESCR, AA.AGDESCR FROM FDOC F "
        + "INNER JOIN fpenc P ON PENUM = F.DREFER LEFT JOIN FPLIN M ON P.PESEQ = M.PESEQ LEFT JOIN FINV I ON I.ISEQ = M.ISEQ LEFT JOIN "
        + "FCLI C ON C.CLISEQ = F.CLISEQ INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 INNER JOIN FAG AA ON AA.AGTNUM = P.PEPAR4 WHERE "
        + "F.DREFER = '" + folio + "' AND F.DITIPMV = 'F' AND F.DCANCELADA = 0 ORDER BY F.DNUM;")

        results = run_query(query) 
        resultado = list(results)

    return render(request, 'detalle_factura.html',{
        'title':folio,
        'resultado':resultado,
        'leyenda':True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def factura_pedido(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        fact_pedido = request.POST.get('fact_pedido', None)

        query = ("SELECT F.DNUM, F.DREFER, I.IDESCR, CAST(M.PLCANT AS SIGNED) AS PLCANT, C.CLINOM, C.CLICONT, CLITEL, CLITEL4, A.AGDESCR, AA.AGDESCR, F.DBRUTO FROM FDOC F "
         "INNER JOIN fpenc P ON PENUM = F.DREFER "
         "LEFT JOIN FPLIN M ON P.PESEQ = M.PESEQ "
         "LEFT JOIN FINV I ON I.ISEQ = M.ISEQ "
         "LEFT JOIN FCLI C ON C.CLISEQ = F.CLISEQ "
         "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
         "INNER JOIN FAG AA ON AA.AGTNUM = P.PEPAR4 "
         f"WHERE F.DREFER = '" + fact_pedido + "';")

        results = run_query(query)
        resultado = list(results)

        total_factura = ("SELECT SUM(F.DBRUTO) FROM FDOC F "
                        +"WHERE F.DREFER = '" + fact_pedido + "';")

        results_factura = run_query(total_factura)
        total_factura = results_factura[0][0] if results_factura else 0

        return render(request, 'factura_pedido.html', {
            'resultado': resultado,
            'leyenda': True,
            'fact_pedido': fact_pedido,
            'total_factura': total_factura,
            'usuario': usuario,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
        return render(request, 'factura_pedido.html', {
            'leyenda': False,
            'usuario': usuario,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

def descargar_excel_activos(request):
    
    query = (f"SELECT PENUM, CLINOM, AGDESCR, PEFECHA, PEPAR3, PEFECHAEMPAQUE, PEDESDE, PEVENCE, PEPAR7 FROM FPLIN L "
                 + "INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG "
                 + "A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
                 + "AND P.PENUM <= 'P99999' AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;")

    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos_activos.xlsx'

    workbook = Workbook()
    worksheet = workbook.active

    encabezados = ["PEDIDO", "RAZON SOCIAL", "VENDEDOR", "FECHA", "CLASIFICACION", "FECHA PARCIALIDAD", "FECHA DE ENTREGA", "NUEVA FECHA"]
    worksheet.append(encabezados)

    # Aplicar formato a los encabezados
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    alternating_colors = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                          PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")]

    for fila in resultado:
        cleaned_row = ["" if (cell == "1900/12/31" or cell == "31/12/1900") else cell for cell in fila]
        worksheet.append(cleaned_row)

    # Aplicar estilos al cuerpo de la tabla (filas alternadas)
    body_font = Font(color="000000")
    body_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=len(encabezados)), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment
            cell.fill = alternating_colors[index % 2]  # Alterna colores de fondo

    # Ajustar márgenes y columnas
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 18
    worksheet.column_dimensions['I'].width = 15

    worksheet.sheet_view.showGridLines = False  # Oculta las líneas de cuadrícula

    # Ajustar los márgenes
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    workbook.save(response)
    return response

@login_required(login_url="login")
def programacion(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if usuario.esvendedor:
        vendedores = CorreosLazzar.objects.filter(esvendedor=usuario.esvendedor).values_list()

        return render(request, 'restringido.html',{
        'title':'Acceso Restringido',
        'vendedores': vendedores,
        'leyenda':False,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
     EMBARQUE = '%EMBARQUE%'

     query = ("SELECT P.PENUM, P.PEVENCE, C.COML5, (SELECT Z.COML2 FROM FCOMENT Z WHERE Z.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2 "
            +"FROM FPENC P "
            +"INNER JOIN FPLIN L ON P.PESEQ = L.PESEQ INNER JOIN FINV I ON L.ISEQ = I.ISEQ "
            +"INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
            +"LEFT JOIN FCOMENT C ON C.COMSEQFACT = CONCAT(10, P.PESEQ) WHERE (L.PLCANT - L.PLSURT) <> 0 "
            +"AND L.PLASIGNADO != 0 AND P.PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' AND P.PENUM <= 'P99999' "
            +"AND P.PESPEDIDO = 1 AND (C.COML5 LIKE '%BORDADO%' OR C.COML5 LIKE '" + EMBARQUE + "') GROUP BY P.PENUM ")

     results = run_query(query)
     resultado = list(results)

     return render(request, 'programacion.html', {
        'resultado': resultado,
        'leyenda': bool(resultado),
        'usuario': usuario,
        'EMBARQUE': EMBARQUE,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })
    
@login_required(login_url="login")
def muestras(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    query = (f"SELECT P.PEFECHA, D.DFECHA, P.PENUM, CLINOM, IFNULL(AGDESCR,'NO TIENE VENDEDOR'), "
            + "IF((P.PEPZAS - P.PEPZASSURT) = 0,'COMPLETO',IF(P.PEPZASSURT > 0, 'PARCIAL','NO SURTIDO')) AS ESTADO, "
            + "FORMAT(IFNULL((SELECT SUM(PLSURT) FROM FPLIN L WHERE L.PESEQ = P.PESEQ GROUP BY P.PENUM),0),0), "
            + "IFNULL(GROUP_CONCAT(D.DNUM),'') AS DOCUMENTOS FROM FPENC P LEFT JOIN FDOC D ON D.DREFER = P.PENUM "
            + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ LEFT JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
            + "WHERE (P.PEPZAS - P.PEPZASSURT) <> 0 AND P.PEFECHA >= '2023-05-05' AND P.PENUM BETWEEN 'M' AND 'M99999' "
            + "AND P.PESPEDIDO = 1 AND (D.DNUM >= 'F' AND D.DNUM <= 'FM9999999' OR D.DNUM IS NULL) GROUP BY P.PENUM;")

    results = run_query(query)

    result_ped = list(results) if len(results) > 0 else []  # Ensure result_ped is always a list = 01/11/2023
    total = len(result_ped)

    return render(request, 'muestras.html', {
        'resultado': results,
        'result_ped': result_ped,
        'total': total,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })


@login_required (login_url="login")
def historial_muestras (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query = (
    f"SELECT P.PEFECHA, D.DFECHA, P.PENUM, DTALON, T.COML4, "
    + "CASE WHEN CLINOM = 'CLIENTE MUESTRAS' THEN(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = CONCAT(10, PESEQ)) ELSE CLINOM END AS CLIENTE, "
    + "IFNULL(AGDESCR,'NO TIENE VENDEDOR'), "
    + "IF((P.PEPZAS - P.PEPZASSURT) = 0, 'COMPLETO', "
    + "IF(P.PEPZASSURT > 0, 'PARCIAL', 'NO SURTIDO')) AS ESTADO, "
    + "FORMAT("
    + "IFNULL("
    + "   (SELECT SUM(PLSURT) FROM FPLIN L WHERE L.PESEQ = P.PESEQ GROUP BY P.PENUM), 0"
    + "), 0) AS PLSURT, "
    + "(SELECT SUM(AICANT) AS 'PIEZAS RECIBIDAS' FROM FDOC Z "
    + "INNER JOIN FAXINV A ON A.DSEQ = Z.DSEQ "
    + "WHERE Z.DNUM >= 'TR' AND DNUM <= 'TR99999' AND AIALMACEN = 00 AND DREFER = P.PENUM), "
    + "IFNULL(GROUP_CONCAT(D.DNUM), '') AS DOCUMENTOS, "
    + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = CONCAT(10, P.PESEQ)) AS COMENT_1 "
    + "FROM FPENC P "
    + "LEFT JOIN FDOC D ON D.DREFER = P.PENUM "
    + "LEFT JOIN FCOMENT T ON T.COMDNUM = D.DNUM "
    + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ "
    + "LEFT JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
    + "WHERE P.PEFECHA BETWEEN '2023-05-05' AND '2024-12-31' "
    + "AND P.PENUM BETWEEN 'M' AND 'M99999' "
    + "AND P.PESPEDIDO = 1 AND (D.DNUM >= 'F' AND D.DNUM <= 'FM9999999' OR D.DNUM IS NULL) "
    + "GROUP BY P.PENUM;"
)


    results = run_query(query) 
                
    if len(results) > 0:
        result_ped = list(results)
        total = len(results)
    else:
        result_ped = 0
        total = 0

    return render(request, 'historial_muestras.html', {
        'resultado': results,
        'result_ped':result_ped,
        'total':total,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })


@login_required(login_url="login")
def muestras_recibidas(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query = (f"SELECT P.PEFECHA, D.DFECHA, P.PENUM, CLINOM, IFNULL(AGDESCR,'NO TIENE VENDEDOR'), "
                    + "IF((P.PEPZAS - P.PEPZASSURT) = 0,'COMPLETO',IF(P.PEPZASSURT > 0, 'PARCIAL','NO SURTIDO')) AS ESTADO, "
                    + "FORMAT(IFNULL((SELECT SUM(PLSURT) FROM FPLIN L WHERE L.PESEQ = P.PESEQ GROUP BY P.PENUM),0),0), "
                    + "IFNULL(GROUP_CONCAT(D.DNUM),'') AS DOCUMENTOS FROM FPENC P LEFT JOIN FDOC D ON D.DREFER = P.PENUM "
                    + "INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ LEFT JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
                    + "WHERE (P.PEPZAS - P.PEPZASSURT) <> 0 AND P.PEFECHA >= '2023-01-01' AND P.PENUM BETWEEN 'M' AND 'M99999' "
                    + "AND P.PESPEDIDO = 1 AND (D.DNUM >= 'F' AND D.DNUM <= 'FM9999999' OR D.DNUM IS NULL) GROUP BY P.PENUM;") 

    results = run_query(query) 
                
    if len(results) > 0:
        result_ped = list(results)
        total = len(results)
    else:
        result_ped = 0
        total = 0

    return render (request, 'muestras_recibidas.html',{
        'resultado': results,
        'result_ped':result_ped,
        'total':total,
        'leyenda': True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required (login_url="login")
def reporte_dia(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
#*******************cantidad vencidos*************************
    query = (f"SELECT PENUM FROM FPLIN L INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ "
                 + "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND "
                 + "P.PENUM >= 'P' AND P.PENUM <= 'P99999' AND datediff(now(), PEDESDE) > 0  AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;")
        
    results = run_query(query) 
        
    if len(results) > 0:
            total_vencidos = len(results)
    else:
            total_vencidos = 0
#**********************cantidad bordados****************************
    fecha_hoy = datetime.datetime.now().date()

    query = (
    f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE, "
    + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10, B.PESEQ)) AS COMENT_1, "
    + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10, B.PESEQ)) AS COMENT_2, "
    + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10, B.PESEQ)) AS COMENT_3, "
    + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10, B.PESEQ)) AS COMENT_4, "
    + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS, "
    + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO, "
    + "(SELECT SUM(L.PLCANT + L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SUMA_TOTAL "
    + "FROM FPENC P "
    + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
    + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
    + "WHERE B.PEPAR0 >= '8' "
    + "AND B.PEPAR0 <= '8ZZZZZZ' "
    + "AND B.PENUM >= 'E' "
    + "AND B.PENUM <= 'E999999999' "
    + "AND B.PEINICIAL = 0 "
    + f"AND B.PEFECHA = '{fecha_hoy}' "  # Agregar la condición de fecha actual
    + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L "
    + "WHERE L.PESEQ = B.PESEQ) > (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);"
)

    results = run_query(query) 

    if len(results) > 0:
        total_bordados = len(results)
    else:
        total_bordados = 0
#******************piezas enviadas*****************
    #enviados_hoy = datetime.datetime.now().date()

    query = (f"SELECT D.DREFER, D.DFECHA, C.CLINOM, D.DNUM, CONCAT('$', FORMAT(DBRUTO, 2)), DTALON, "
                    + "(SELECT COML4 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM) AS GUIA,  " 
                    + "(SELECT A.AGDESCR FROM FAG A WHERE A.AGTNUM = D.DPAR1) AS VENDEDOR,"
                    + "(SELECT FORMAT(SUM(N.AICANTF), 0) FROM FAXINV N WHERE N.DSEQ = D.DSEQ) "
                    + "FROM FDOC D "
                    + "INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                    + "WHERE D.DFECHA = DATE(NOW()) AND D.DNUM >= 'F0' AND D.DNUM <= 'FE99999999' "
                    + "AND D.DCANCELADA = 0;") 
    
    results = run_query(query) 

    if len(results) > 0:
        total_enviados = len(results)
    else:
        total_enviados = 0
#***********cantidad factura**********************
    query = ("SELECT SUM(F.DBRUTO) AS suma_dbruto "
        "FROM FDOC F "
        "INNER JOIN fpenc P ON PENUM = F.DREFER "
        "LEFT JOIN FPLIN M ON P.PESEQ = M.PESEQ "
        "LEFT JOIN FINV I ON I.ISEQ = M.ISEQ "
        "LEFT JOIN FCLI C ON C.CLISEQ = F.CLISEQ "
        "INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
        "INNER JOIN FAG AA ON AA.AGTNUM = P.PEPAR4 "
        "WHERE F.DFECHA = DATE(NOW()) "
        "AND F.DITIPMV = 'F' "
        "AND F.DCANCELADA = 0;") 
    results = run_query(query) 

    if any(result[0] is None for result in results):
        suma_dbruto = 0
    else:
        suma_dbruto = sum(result[0] for result in results)
        
    formatted_amount = "${:,.2f}".format(suma_dbruto) # Imprime en formato al $uma

    return render(request, 'reporte_dia.html',{
        'total_vencidos':total_vencidos,
        'total_bordados':total_bordados,
        'total_enviados':total_enviados,
        'suma_dbruto':formatted_amount,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })
    
@login_required(login_url='login')
def search_pedido(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        encontrar_pedido = request.POST.get('encontrar_pedido', None)

        query = (f"SELECT PENUM, CLINOM, AGDESCR, PEFECHA, PEPAR3, PEFECHAEMPAQUE, PEDESDE, PEVENCE, PEPAR7 FROM FPLIN L "
                + "INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG "
                + "A ON A.AGTNUM = P.PEPAR1 WHERE PEFECHA >= '2023-01-01' "
                + f"AND P.PENUM = '{encontrar_pedido}' AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;") 

        results = run_query(query)
        resultado = list(results)

        return render(request, 'search_pedido.html', {
            'resultado': resultado,
            'leyenda': True,
            'encontrar_pedido': encontrar_pedido,
            'usuario': usuario.esvendedor,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render(request, 'search_pedido.html', {
            'leyenda': False,
            'usuario': usuario.esvendedor,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

#Botones de actualización -------ACTIVOS------ 17/nov/23
def boton_activos(request):
    query = (f"SELECT PENUM, CLINOM, AGDESCR, PEFECHA, PEPAR3, PEFECHAEMPAQUE, PEDESDE, PEVENCE, PEPAR7 FROM FPLIN L "
                 + "INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ INNER JOIN FAG "
                 + "A ON A.AGTNUM = P.PEPAR1 WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' "
                 + "AND P.PENUM <= 'P99999' AND PESPEDIDO = 1 GROUP BY PENUM ORDER BY PENUM;") 
    
    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="data.txt"'
    response.write(resultado)

    return response
    
@login_required(login_url='login')
def info(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    mexico_timezone = pytz.timezone('America/Mexico_City')
    fecha_actual = datetime.now(mexico_timezone).strftime('%Y-%m-%d')
    fecha_mes = (datetime.now(mexico_timezone) - timedelta(days=31)).strftime('%Y-%m-%d')
    fecha_año = (datetime.now(mexico_timezone) - timedelta(days=365)).strftime('%Y-%m-%d')
    total_usuarios = User.objects.count()
    primera_fecha = ''
    fecha_final = ''

    if request.method == 'POST':
        primera_fecha = request.POST['primera_fecha']
        fecha_final = request.POST['fecha_final']
    
        query = ( #PEDIDOS DEL DIA
            f"SELECT COUNT(PENUM) AS TOTALPEDIDOS FROM FPENC "
            f"WHERE PENUM >= 'P' AND PENUM <= 'P99999' AND PESPEDIDO = 1 AND PEFECHA BETWEEN '{primera_fecha}' AND '{fecha_final}' "
        )

        results = run_query(query)
        resultado = list(results)

        query_mes = ( #PEDIDOS DEL MES
            f"SELECT COUNT(PENUM) AS TOTALPEDIDOS FROM FPENC "
            f"WHERE PENUM >= 'P' AND PENUM <= 'P99999' AND PESPEDIDO = 1 AND PEFECHA BETWEEN '{primera_fecha}' AND '{fecha_final}' "
        )

        results_mes = run_query(query_mes)
        resultado_mes = list(results_mes)

        query_ingresos = ( #INGRESOS DEL DIA
                        f"SELECT CONCAT('$', FORMAT(SUM(DBRUTO), 2)) "
                        + "FROM FDOC D "
                        + "INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                        + "WHERE D.DREFER >= 'P' AND D.DREFER <= 'P99999' AND D.DFECHA BETWEEN '" + primera_fecha + "' AND '" + fecha_final + "' AND D.DNUM >= 'F' AND D.DNUM <= 'F99999999' "
                        + "AND D.DCANCELADA = 0; "
                        ) 

        results_ingresos = run_query(query_ingresos)
        resultado_ingresos = list(results_ingresos) 

        query_embarque_now = ( #PEDIDOS EMBARCADOS
                        f"SELECT COUNT(D.DREFER) FROM FDOC D "
                        + "INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                        + "LEFT JOIN FCOMENT C4 ON C4.COMDNUM = D.DNUM "
                        + "WHERE D.DREFER >= 'P' AND D.DREFER <= 'P99999' "
                        + f"AND D.DFECHA BETWEEN '{primera_fecha}' AND '{fecha_final}' "
                        + "AND D.DNUM >= 'F' AND D.DNUM <= 'FE99999999' "
                        + f"AND D.DCANCELADA = 0 "
                        + "AND (DTALON LIKE 'DHL' OR DTALON LIKE 'LOCAL' OR DTALON LIKE 'PAQUETEXPRESS' OR DTALON LIKE 'CLIENTE RECOGE' OR D.DTALON LIKE 'CHAVA'); "
                        )

        results_embarque_now = run_query(query_embarque_now) 
        resultado_embarque_now = list(results_embarque_now)

        query_piezas_envios = ( #CANT PIEZAS ENVIADAS
                        f"SELECT SUM(A.AICANTF) FROM FAXINV A "
                        + "INNER JOIN FDOC D ON A.DSEQ = D.DSEQ "
                        + "LEFT JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                        + "LEFT JOIN FCOMENT C4 ON C4.COMDNUM = D.DNUM "
                        + "WHERE D.DREFER >= 'P' AND D.DREFER <= 'P99999' "
                        + f"AND D.DFECHA BETWEEN '{primera_fecha}' AND '{fecha_final}' "
                        + "AND D.DNUM >= 'F' AND D.DNUM <= 'FE99999999' "
                        + "AND D.DCANCELADA = 0 "
                        + "AND (D.DTALON LIKE 'DHL' OR D.DTALON LIKE 'LOCAL' OR D.DTALON LIKE 'PAQUETEXPRESS' OR D.DTALON LIKE 'CLIENTE RECOGE' OR D.DTALON LIKE 'CHAVA');"
                        )

        results_piezas_envios = run_query(query_piezas_envios) 
        resultado_piezas_envios = list(results_piezas_envios)
        # query_compradores = (
        #     f"SELECT C.CLISEQ, C.CLINOM, COUNT(P.PENUM) AS NUMERO_PIEZAS FROM FCLI C "
        #     "INNER JOIN FPENC P ON P.PESEQ = C.PESEQ "
        #     "GROUP BY C.CLISEQ, C.CLINOM "
        #     "ORDER BY NUMERO_PIEZAS "
        #     "LIMIT 10; "
        # )
        return render(request, 'info.html', {
            'resultado': resultado,
            'resultado_mes': resultado_mes, 
            'resultado_ingresos': resultado_ingresos,
            'resultado_embarque_now': resultado_embarque_now,
            'resultado_piezas_envios': resultado_piezas_envios,
            'total_usuarios': total_usuarios,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
            'leyenda': True,
        })
    else:
        return render(request, 'info.html', {
            'ventas': ventas,
            'acceso_total': acceso_total,
            'mesa_control': mesa_control,
            'leyenda': False,
        })
@login_required(login_url="login")
def historial_programacion (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    EMBARQUE = '%EMBARQUE%'

    query = ("SELECT P.PENUM, P.PEVENCE, C.COML5, (SELECT Z.COML2 FROM FCOMENT Z WHERE Z.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2 "
            +"FROM FPENC P "
            +"INNER JOIN FPLIN L ON P.PESEQ = L.PESEQ INNER JOIN FINV I ON L.ISEQ = I.ISEQ "
            +"INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
            +"LEFT JOIN FCOMENT C ON C.COMSEQFACT = CONCAT(10, P.PESEQ) WHERE "
            +"P.PEFECHA BETWEEN '2023-01-01' AND '2023-12-31' AND P.PENUM BETWEEN 'P' AND 'P99999' "
            +"AND P.PESPEDIDO = 1 AND (C.COML5 LIKE '%BORDADO%' OR C.COML5 LIKE '" + EMBARQUE + "') GROUP BY P.PENUM ")

    results = run_query(query)
    resultado = list(results)

    return render(request, 'historial_programacion.html', {
        'resultado': resultado,
        'leyenda':bool(resultado),
        'EMBARQUE': EMBARQUE,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def reporte_activos (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query = ("SELECT PENUM AS PEDIDO, CLINOM AS CLIENTE, AGDESCR AS VENDEDOR, PEFECHA AS FECHA_INGRESO, PEPAR3 AS CLASIFICACION, "
            +"CONCAT('$ ', FORMAT(PEBRUTO, 2)), PEDESDE AS FECHA_ENTREGA, PEVENCE AS NUEVA_FECHA_ENTREGA FROM FPLIN L "
            +"INNER JOIN FPENC P ON P.PESEQ = L.PESEQ INNER JOIN FCLI C ON C.CLISEQ = P.CLISEQ  INNER JOIN FAG A ON A.AGTNUM = P.PEPAR1 "
            +"WHERE (PLCANT-PLSURT) <> 0 AND PEFECHA >= '2023-01-01' AND P.PENUM >= 'P' AND P.PENUM <= 'P99999' AND PESPEDIDO = 1 "
            +"GROUP BY PENUM ORDER BY PENUM;")

    results = run_query(query)
    resultado = list(results)

    return render(request, 'reporte_activos.html', {
        'resultado': resultado,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def reporte_facturados (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    query = ("SELECT P.PENUM AS PEDIDO, P.PEPAR3 AS CLASIFICACIÓN, P.PEFECHA AS FECHA_INGRESO, P.PEDESDE AS DESDE, P.PEVENCE AS VENCE, "
            +"D.DNUM AS FACTURA, D.DFECHA AS FECHA_FACTURA FROM FPENC P LEFT JOIN FDOC D ON D.DREFER = P.PENUM AND D.DNUM >= 'F0' "
            +"AND D.DNUM <= 'FE99999' AND D.DCANCELADA = 0 WHERE P.PENUM >= 'P' AND P.PENUM <= 'P999999' AND P.PEFECHA >= '2023-01-01' "
            +"AND P.PEFECHA <= '" + fecha_actual + "' AND P.PESPEDIDO = 1;")

    results = run_query(query)
    resultado = list(results)

    return render(request, 'reporte_facturados.html', {
        'resultado': resultado,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })