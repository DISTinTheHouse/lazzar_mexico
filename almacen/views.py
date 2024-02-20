from django.shortcuts import render
from django.contrib.auth.decorators import login_required
import MySQLdb
from .models import codigo_almacen
from mainapp.models import CorreosLazzar
from django.db import connection
#EXCEL
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse

DB_HOST = 'as.galio.net' 
DB_USER = 'u129' 
DB_PASS = 'lbrfyg' 
DB_NAME = 'db129lazzar'
DB_PORT = 61129

def run_query(query=''): 
    datos = [DB_HOST, DB_USER, DB_PASS, DB_NAME, DB_PORT] 
    
    conn = MySQLdb.connect(*datos) # Conectar a la base de datos 
    cursor = conn.cursor()         # Crear un cursor 
    cursor.execute(query)          # Ejecutar una consulta 

    if query.upper().startswith('SELECT'): 
        data = cursor.fetchall()   # Traer los resultados de un select 
    else: 
        conn.commit()              # Hacer efectiva la escritura de datos 
        data = None 
    
    cursor.close()                 # Cerrar el cursor 
    conn.close()                   # Cerrar la conexión 

    return data

# Create your views here.

@login_required(login_url="login")
def pedidos_index(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    
    
    if usuario.esvendedor:
        vendedores = CorreosLazzar.objects.filter(esvendedor=usuario.esvendedor).values_list()

        return render(request, 'restringido.html',{
        'title':'Acceso Restringido',
        'vendedores': vendedores,
        'leyenda':False,
        })
    else:
        return render(request, 'pedidos_index.html',{
        'leyenda':False
        })

@login_required(login_url="login")
def materiales(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    
    if usuario.esvendedor:
        vendedores = CorreosLazzar.objects.filter(esvendedor=usuario.esvendedor).values_list()

        return render(request, 'restringido.html',{
        'title':'Acceso Restringido',
        'vendedores': vendedores,
        'leyenda':False,
        })
    else:
        return render(request, 'materiales.html',{
        'leyenda':False,
        })

#se agrego stock telas - Jesús - 11/09/23
@login_required(login_url="login")
def telas(request):

    query_telas = (f"SELECT ICOD, IDESCR, CAJSERIE, FORMAT(CAJCANT,2), CAJOBS FROM fcajas C "
             +"INNER JOIN FINV I ON I.ISEQ = C.ISEQ "
             +"WHERE CAJALM = '04' AND CAJFACTURA = '';")
    
    results = run_query(query_telas)
    resultados_telas = list(results)

    return render(request, 'telas_stock.html',{
        'leyenda': True,
        'resultado_telas': resultados_telas,
    })

#Se agrego models.py almacen y se configuro form y pagina - Jesús - 12/09/23
@login_required(login_url="login")
def stock_alm(request):
    almacenes = codigo_almacen.objects.all()
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        cod_almacen = request.POST.get('almacen')
        almacen = codigo_almacen.objects.get(cod_almacen=cod_almacen)  
        resultados = []
        resultados_telas = []
        resultados_prod_terminado = []
        resultados_eder = []

        if cod_almacen == '02':

            query = (f"SELECT ICOD as CODIGO, IDESCR as DESCRIPCION, ROUND(SUM(A.ALMCANT-A.ALMASIGNADO), 0) as INVENTARIO "
                +"FROM FINV I INNER JOIN FALM A ON A.ISEQ = I.ISEQ "
                +"WHERE A.ALMNUM = '" + cod_almacen + "' AND A.ALMCANT <> '0' AND A.ALMCANT <> 0 GROUP BY ICOD ;")

            results = run_query(query)
            resultados = list(results)

        if cod_almacen == '03' or '04' or '07' or '11' or '15':
            query_telas = (f"SELECT ICOD, IDESCR, CAJSERIE, FORMAT(CAJCANT,2), CAJOBS FROM fcajas C "
                +"INNER JOIN FINV I ON I.ISEQ = C.ISEQ "
                +"WHERE CAJALM = '" + cod_almacen + "' AND CAJFACTURA = '';")
        
            results = run_query(query_telas)
            resultados_telas = list(results)
        
            if cod_almacen == '47': 
                    query_eder = (f"SELECT ICOD as CODIGO,IDESCR as DESCRIPCION, ROUND(SUM(A.ALMCANT-A.ALMASIGNADO),0 ) as INVENTARIO FROM FINV I INNER JOIN FALM A ON A.ISEQ = I.ISEQ "
                                        +"WHERE ICOD >= 'A' AND ICOD <= 'Z99999999' AND A.ALMNUM = '47' AND A.ALMCANT <> '0' AND A.ALMCANT <> 0 GROUP BY ICOD ")

                    results = run_query(query_eder)
                    resultados_eder = list(results)
        
            if cod_almacen == '00': 
                query_prod_terminado = (f"SELECT ICOD as CODIGO,IDESCR as DESCRIPCION, ROUND(SUM(A.ALMCANT-A.ALMASIGNADO),0 ) as INVENTARIO FROM FINV I INNER JOIN FALM A ON A.ISEQ = I.ISEQ "
                                    +"WHERE ICOD >= 'A' AND ICOD <= 'Z99999999' AND A.ALMNUM = '00' AND A.ALMCANT <> '0' AND A.ALMCANT <> 0 GROUP BY ICOD ")

                results = run_query(query_prod_terminado)
                resultados_prod_terminado = list(results)

                return render(request, "stock_alm.html", {
                        'almacenes': almacenes,
                        'leyenda': True,
                        'resultados_prod_terminado': resultados_prod_terminado,
                        'prod_term': bool(resultados_prod_terminado),
                        'resultados_eder':resultados_eder,
                        'eder':bool(resultados_eder),
                        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
                        'acceso_total': acceso_total,
                        'mesa_control': mesa_control, # End Admin-Groups
                        })

        return render(request, 'stock_alm.html', {
            'almacenes': almacenes,
            'almacen': almacen,  # Pasa el almacén seleccionado a la plantilla
            'resultados': resultados,
            'resultados_telas': resultados_telas,
            'resultados_prod_terminado': resultados_prod_terminado,
            'resultados_eder':resultados_eder,
            'avios': bool(resultados),
            'telas': bool(resultados_telas),
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render(request, "stock_alm.html", {
            'almacenes': almacenes,
            'leyenda': False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required(login_url="login") 
def descargar_excel_04(request):
    
    query = (f"SELECT ICOD, IDESCR, CAJSERIE, FORMAT(CAJCANT,2), CAJOBS FROM fcajas C "
                +"INNER JOIN FINV I ON I.ISEQ = C.ISEQ "
                +"WHERE CAJALM = '04' AND CAJFACTURA = '';")

    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_almacen_04.xlsx'

    workbook = Workbook()
    worksheet = workbook.active

    encabezados = ["CODIGO", "DESCRIPCIÓN", "FOLIO ROLLO", "METROS", "COMENTARIO TELA"]
    worksheet.append(encabezados)

    # Aplicar formato a los encabezados
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    alternating_colors = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                          PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")]

    for fila in resultado:
        cleaned_row = ["" if (cell == "1900/12/31" or cell == "31/12/1900") else cell for cell in fila]
        worksheet.append(cleaned_row)

    # Aplicar estilos al cuerpo de la tabla (filas alternadas)
    body_font = Font(color="000000")
    body_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=len(encabezados)), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment
            cell.fill = alternating_colors[index % 2]  # Alterna colores de fondo

    # Ajustar márgenes y columnas
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 20
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 25

    worksheet.sheet_view.showGridLines = False  # Oculta las líneas de cuadrícula

    # Ajustar los márgenes
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    workbook.save(response)
    return response

@login_required(login_url="login")
def pedidos_q(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 20/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        inic_fech = request.POST['inic_fech']
        fin_fech = request.POST['fin_fech']

        query = (f"SELECT P.PEFECHA , P.PENUM , P.PEDESDE , P.PEVENCE , N.PRVNOM , S.ICOD , S.IDESCR , SUM(C.PLCANT) from FPENC P "
		        +"INNER JOIN FPLIN C ON C.PESEQ = P.PESEQ "
		        +"INNER JOIN FINV S ON S.ISEQ = C.ISEQ "
		        +"INNER JOIN FCOMENT T ON T.COMSEQFACT = CONCAT(10, P.PESEQ) "
		        +"INNER JOIN FPRV N ON N.PRVCOD = T.COML1 "
                +"WHERE P.PEFECHA BETWEEN '" + inic_fech + "' AND '" + fin_fech + "' AND P.PENUM >= 'Q' AND P.PENUM <= 'Q99999' AND P.PESPEDIDO = 1 "
                +"GROUP BY P.PENUM ")
                
        
        results = run_query(query)
        resultado = list(results)


        return render(request, 'pedidos_q.html',{
            'leyenda': True,
            'fech_inic':inic_fech,
            'fecha_fin': fin_fech, 
            'resultado': resultado,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 20/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render(request, 'pedidos_q.html',{
            'leyenda': False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 20/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required(login_url="login") 
def descargar_excel_avios(request):
    
    query = (f"SELECT ICOD as CODIGO, IDESCR as DESCRIPCION, ROUND(SUM(A.ALMCANT-A.ALMASIGNADO), 0) as INVENTARIO "
                +"FROM FINV I INNER JOIN FALM A ON A.ISEQ = I.ISEQ "
                +"WHERE A.ALMNUM = '02' AND A.ALMCANT <> '0' AND A.ALMCANT <> 0 GROUP BY ICOD ;")

    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=stock_almacen_02_avíos.xlsx'

    workbook = Workbook()
    worksheet = workbook.active

    encabezados = ["CODIGO", "DESCRIPCIÓN", "INVENTARIO"]
    worksheet.append(encabezados)

    # Aplicar formato a los encabezados
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    alternating_colors = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                          PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")]

    for fila in resultado:
        cleaned_row = ["" if (cell == "1900/12/31" or cell == "31/12/1900") else cell for cell in fila]
        worksheet.append(cleaned_row)

    # Aplicar estilos al cuerpo de la tabla (filas alternadas)
    body_font = Font(color="000000")
    body_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=len(encabezados)), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment
            cell.fill = alternating_colors[index % 2]  # Alterna colores de fondo

    # Ajustar márgenes y columnas
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 40
    worksheet.column_dimensions['D'].width = 30

    worksheet.sheet_view.showGridLines = False  # Oculta las líneas de cuadrícula

    # Ajustar los márgenes
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    workbook.save(response)
    return response

@login_required(login_url="login")
def ordenes_e (request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        inic_fech = request.POST['inic_fech']
        fin_fech = request.POST['fin_fech']

        query = (f"SELECT P.PEFECHA, P.PENUM, P.PEOTROTXT, P.PEDESDE , P.PEVENCE, C.PRVNOM, I.ICOD, I.IDESCR, SUM(L.PLCANT) FROM FPLIN L "
                 +"INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
                 +"INNER JOIN FPRV C ON C.PRVSEQ = P.PRVSEQ "
                 +"INNER JOIN FINV I ON I.ISEQ = L.ISEQ "
                 +"WHERE P.PEFECHA BETWEEN '" + inic_fech + "' AND '" + fin_fech + "' AND P.PENUM >= 'E' AND P.PENUM <= 'E99999' "
                 +"AND P.PEPAR0 >= '6' AND P.PEPAR0 <= '6ZZZZZZ' AND P.PEOTROTXT >= 'Q' AND P.PEOTROTXT <= 'Q99999' "
                 +"GROUP BY P.PENUM ;")
        
        results = run_query(query)
        respuesta = list(results)

        return render (request, 'ordenes_e.html',{
            'leyenda':True,
            'fecha_inic':inic_fech,
            'fecha_fin':fin_fech,
            'respuesta':respuesta,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render(request, 'ordenes_e.html',{
                'leyenda':False,
                'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
                'acceso_total': acceso_total,
                'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required (login_url="login")
def detalle_ordenes_e(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        folio = request.POST['variable']
    else:
        folio = 'detboton'

    query = (f"SELECT P.PEFECHA, P.PENUM, S.ICOD, S.IDESCR, C.PLCANT FROM FPENC P "
                 +"INNER JOIN FPLIN C ON C.PESEQ = P.PESEQ "
                 +"INNER JOIN FINV S ON S.ISEQ = C.ISEQ "
                 +"WHERE P.PENUM = '" + folio + "' AND P.PENUM >= 'E' AND P.PENUM <= 'E99999' ")
    
    results = run_query(query)
    respuesta = list(results)

    total_query = (f"SELECT P.PENUM, SUM(C.PLCANT) FROM FPENC P "
                 +"INNER JOIN FPLIN C ON C.PESEQ = P.PESEQ "
                 +"WHERE P.PENUM = '" + folio + "' AND P.PENUM >= 'E' AND P.PENUM <= 'E99999' ")
    
    total_detalle = run_query(total_query)
    totales_detalle = list(total_detalle)

    return render (request, 'detalle_ordenes_e.html',{
        'title':folio,
        'leyenda':True,
        'respuesta':respuesta,
        'totales_detalle':totales_detalle,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def pedidos_pq(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        inicio_fech = request.POST['inicio_fech']
        fin_fecha = request.POST['fin_fecha']

        query = (f"SELECT D.DREFER, D.DFECHA, C.CLINOM, D.DNUM, CONCAT('$', FORMAT(DBRUTO, 2)), DTALON, " 
                + "(SELECT COML4 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM) AS GUIA, "
                + "(SELECT A.AGDESCR FROM FAG A WHERE A.AGTNUM = D.DPAR1) AS VENDEDOR, "
                + "(SELECT FORMAT(SUM(N.AICANTF), 0) FROM FAXINV N WHERE N.DSEQ = D.DSEQ) AS CANTIDAD_PIEZAS, "
                + "C5.COML5 AS COMENTARIO " #SE AGREGO COMENTARIO DE OBSERVACIONES
                + "FROM FDOC D "
                + "INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                + "LEFT JOIN FCOMENT C4 ON C4.COMDNUM = D.DNUM "
                + "LEFT JOIN FCOMENT C5 ON C5.COMDNUM = D.DNUM "
                + "WHERE D.DREFER >= 'PQ' AND D.DREFER <= 'PQ99999' AND D.DFECHA BETWEEN '" + inicio_fech + "' AND '" + fin_fecha + "' AND D.DNUM >= 'FT' AND D.DNUM <= 'FT99999999' "
                + "AND D.DCANCELADA = 0; ")
        
        # query = (f"SELECT D.DREFER, D.DFECHA, C.CLINOM, D.DNUM, CONCAT('$', FORMAT(DBRUTO, 2)), DTALON, "  
        #         + "(SELECT COML4 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM) AS GUIA, "
        #         + "(SELECT A.AGDESCR FROM FAG A WHERE A.AGTNUM = D.DPAR1) AS VENDEDOR, "
        #         + "(SELECT FORMAT(SUM(N.AICANTF), 0) FROM FAXINV N WHERE N.DSEQ = D.DSEQ) AS CANTIDAD_PIEZAS, "
        #         + "(SELECT COML5 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM) AS COMENTARIO "             VERSION ANTERIOR PERO SIN OPTIMIZAR CON LEFT JOIN
        #         + "FROM FDOC D "
        #         + "INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
        #         + "WHERE D.DREFER >= 'PQ' AND D.DREFER <= 'PQ99999' AND D.DFECHA BETWEEN '" + inicio_fech + "' AND '" + fin_fecha + "' AND D.DNUM >= 'FT' AND D.DNUM <= 'FT99999999' "
        #         + "AND D.DCANCELADA = 0; ")


        results = run_query(query)

        results_query = list(results) if len (results) > 0 else []
        total = len(results_query)

        return render(request, 'pedidos_pq.html',{
            'leyenda':True,
            'inicio_fech':inicio_fech,
            'fin_fecha':fin_fecha,
            'resultado':results,
            'total':total,
            'results_query': results_query,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    else:
        return render (request, 'pedidos_pq.html',{
            'leyenda':False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
        
@login_required(login_url='login')  
def search_pq(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
            encontrar_pq = request.POST.get('encontrar_pq', None)

            query = (
                    f"SELECT D.DREFER, D.DFECHA, C.CLINOM, D.DNUM, CONCAT('$', FORMAT(DBRUTO, 2)), DTALON, "
                    f"(SELECT COML4 FROM FCOMENT C WHERE C.COMDNUM = D.DNUM) AS GUIA, "
                    f"(SELECT A.AGDESCR FROM FAG A WHERE A.AGTNUM = D.DPAR1) AS VENDEDOR, "
                    f"(SELECT FORMAT(SUM(N.AICANTF), 0) FROM FAXINV N WHERE N.DSEQ = D.DSEQ) AS CANTIDAD_PIEZAS, "
                    f"C5.COML5 AS COMENTARIO "
                    f"FROM FDOC D "
                    f"INNER JOIN FCLI C ON C.CLISEQ = D.CLISEQ "
                    f"LEFT JOIN FCOMENT C4 ON C4.COMDNUM = D.DNUM "
                    f"LEFT JOIN FCOMENT C5 ON C5.COMDNUM = D.DNUM "
                    f"WHERE D.DREFER = '{encontrar_pq}' AND D.DNUM >= 'FT' AND D.DNUM <= 'FT99999999' "
                    f"AND D.DCANCELADA = 0;"
                )

            
            results = run_query(query)
            resultado = list(results)

            return render(request, 'buscar_pq.html', {
                'encontrar_pq':encontrar_pq,
                'leyenda':True,
                'resultado':resultado,
                'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
                'acceso_total': acceso_total,
                'mesa_control': mesa_control, # End Admin-Groups
            })        
            
@login_required(login_url='login')
def search_q(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        encontrar_p = request.POST.get('encontrar_p', None)

        query = (f"SELECT P.PEFECHA , P.PENUM , P.PEDESDE , P.PEVENCE , N.PRVNOM , S.ICOD , S.IDESCR , C.PLCANT from FPENC P "
		        +"INNER JOIN FPLIN C ON C.PESEQ = P.PESEQ "
		        +"INNER JOIN FINV S ON S.ISEQ = C.ISEQ "
		        +"INNER JOIN FCOMENT T ON T.COMSEQFACT = CONCAT(10, P.PESEQ) "
		        +"INNER JOIN FPRV N ON N.PRVCOD = T.COML1 "
                +"WHERE P.PENUM = '" + encontrar_p + "' ")
        results = run_query(query)
        resultado = list(results)

        return render(request, 'search_q.html', {
            'resultado': resultado,
            'leyenda': True,
            'encontrar_p': encontrar_p,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
@login_required(login_url="login")
def search_e(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        encontrar_e = request.POST.get('encontrar_e', None)

        query = (f"SELECT P.PEFECHA, P.PENUM, P.PEDESDE , P.PEVENCE, C.PRVNOM, I.ICOD, I.IDESCR, SUM(L.PLCANT) FROM FPLIN L "
                 +"INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
                 +"INNER JOIN FPRV C ON C.PRVSEQ = P.PRVSEQ "
                 +"INNER JOIN FINV I ON I.ISEQ = L.ISEQ "
                 +"WHERE P.PENUM = '" + encontrar_e + "' "
                 +"AND P.PEPAR0 >= '6' AND P.PEPAR0 <= '6ZZZZZZ' "
                 +"GROUP BY P.PENUM ;")
        
        results = run_query(query)
        resultado = list(results)

        return render(request, 'search_e.html', {
            'resultado': resultado,
            'leyenda': True,
            'encontrar_e': encontrar_e,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
        
@login_required(login_url="login")
def orden_e_vencido(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 29/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query_orden_e = (
    "SELECT P.PEFECHA, P.PENUM, P.PEOTROTXT, P.PEDESDE , P.PEVENCE, C.PRVNOM, "
    "SUBSTRING_INDEX(Z.COML5, 'INCLUYE', -1) AS comentario, "
    "SUM(L.PLCANT) "
    "FROM FPLIN L "
    "INNER JOIN FPENC P ON P.PESEQ = L.PESEQ "
    "INNER JOIN FPRV C ON C.PRVSEQ = P.PRVSEQ "
    "INNER JOIN FINV I ON I.ISEQ = L.ISEQ "
    "LEFT JOIN FCOMENT Z ON Z.COMSEQFACT = CONCAT(10, P.PESEQ) "
    "WHERE DATEDIFF(NOW(), P.PEDESDE) > 0 AND P.PENUM >= 'E' AND P.PENUM <= 'E99999' "
    "AND P.PEPAR0 >= '6' AND P.PEPAR0 <= '6ZZZZZZ' AND P.PEOTROTXT >= 'P' AND P.PEOTROTXT <= 'P99999' AND PEINICIAL = 0 "
    "GROUP BY P.PENUM ;"
    )
        
    results_orden_e = run_query(query_orden_e)
    resultado_orden_e = list(results_orden_e)

    return render(request, 'vencidos_orden_e.html',{
        'vencidos_orden_e': resultado_orden_e,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 29/12/23
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })