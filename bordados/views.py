from django.shortcuts import render, HttpResponse, redirect
from django.db.models import F, Q
from django.contrib.auth.decorators import login_required
from mainapp.models import CorreosLazzar
import datetime
#excel
from openpyxl.worksheet.page import PageMargins
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.http import HttpResponse

import MySQLdb

DB_HOST = 'as.galio.net' 
DB_USER = 'u129' 
DB_PASS = 'lbrfyg' 
DB_NAME = 'db129lazzar'
DB_PORT = 61129

def run_query(query=''): 
    datos = [DB_HOST, DB_USER, DB_PASS, DB_NAME, DB_PORT] 
    
    conn = MySQLdb.connect(*datos) # Conectar a la base de datos 
    cursor = conn.cursor()         # Crear un cursor 
    cursor.execute(query)          # Ejecutar una consulta 

    if query.upper().startswith('SELECT'): 
        data = cursor.fetchall()   # Traer los resultados de un select 
    else: 
        conn.commit()              # Hacer efectiva la escritura de datos 
        data = None 
    
    cursor.close()                 # Cerrar el cursor 
    conn.close()                   # Cerrar la conexión 

    return data

def pedidos(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    
    if usuario.esvendedor:
        vendedores = CorreosLazzar.objects.filter(esvendedor=usuario.esvendedor).values_list()
        
        return render(request, 'restringido.html', {
            'title': 'Acceso Restringido',
            'vendedores': vendedores,
            'leyenda': False,
        })
    else:
        return render(request, 'pedidos.html', {
            'leyenda': False,
        })


@login_required(login_url="login")
def pendientes_folios(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    
    query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE, "
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PEPAR0 >= '8' "
                    + "AND B.PEPAR0 <= '8ZZZZZZ' "
                    + "AND B.PENUM >= 'E' "
                    + "AND B.PENUM <= 'E999999999' "
                    + "AND B.PEINICIAL = 0 "
                    + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L "
                    + "WHERE L.PESEQ = B.PESEQ) > (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

    results = run_query(query) 

    resultado = list(results)

    return render(request, 'pendientes.html',{
            'title':'Listado de Pedidos',
            'resultado':resultado,
            'leyenda':True,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def detalle_bordados(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        folio = request.POST['variable']
    else:
        folio = 'detboton'

    query = (
        f"SELECT P.PENUM, P.PEFECHA, P.PEDESDE, P.PEVENCE,"
        + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_1,"
        + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_2,"
        + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_3,"
        + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,P.PESEQ)) AS COMENT_4 "
        + "FROM FPENC P "
        + "WHERE P.PENUM = '" + folio + "' ")

    results = run_query(query)
    segresultado = list(results)

    largo = len(segresultado)
    i = 1
    consulta = ""

    if largo > 0:
        for orden in segresultado:
            if i == largo:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "'"
                i += 1
            else:
                consulta = consulta + "P.PENUM = '" + orden[0].strip() + "' OR "
                i += 1

        detalle_bord = (
            f"SELECT P.PENUM, I.ICOD, I.IDESCR, L.PLCANT, L.PLSURT, (L.PLCANT-L.PLSURT) FROM FPENC P "
            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
            + "(" + consulta + ") ORDER BY P.PENUM;"
        )

        ops_bord = run_query(detalle_bord)
        detalle_bordados = list(ops_bord)

        total_query_bor = (
            f"SELECT P.PENUM, SUM(L.PLCANT), SUM(L.PLSURT), SUM((L.PLCANT-L.PLSURT)) FROM FPENC P "
            + "INNER JOIN FPLIN L ON L.PESEQ = P.PESEQ INNER JOIN FINV I ON I.ISEQ = L.ISEQ WHERE "
            + "(" + consulta + ") GROUP BY P.PENUM;"
        )

        total_bor = run_query(total_query_bor)
        totales_bor = list(total_bor)
    else:
        detalle_bordados = None
        totales_bor = None

    return render(request, 'detalle_bordados.html', {
        'title': folio,
        'leyenda': True,
        'segresultado': segresultado,
        'detalle_bordados': detalle_bordados,
        'totales_bor': totales_bor,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })
 

@login_required(login_url="login") 
def descargar_excel(request):
    query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE, "
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PEPAR0 >= '8' "
                    + "AND B.PEPAR0 <= '8ZZZZZZ' "
                    + "AND B.PENUM >= 'E' "
                    + "AND B.PENUM <= 'E999999999' "
                    + "AND B.PEINICIAL = 0 "
                    + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L "
                    + "WHERE L.PESEQ = B.PESEQ) > (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos_pendientes.xlsx'

    workbook = Workbook()
    worksheet = workbook.active

    encabezados = ["PEDIDO", "FECHA", "ORDEN", "PROVEEDOR", "FECHA DE SURTIDO", "RECEPCIÓN DE LA ORDEN", "FECHA DE ENTREGA", "SURTIDO", "ESTATUS", "RECIBIÓ", "DESEBRO", "PIEZAS", "ENTREGADO"]
    worksheet.append(encabezados)

    # Aplicar formato a los encabezados
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    alternating_colors = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                          PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")]

    for fila in resultado:
        cleaned_row = ["" if (cell == "1900/12/31" or cell == "31/12/1900") else cell for cell in fila]
        worksheet.append(cleaned_row)

    # Aplicar estilos al cuerpo de la tabla (filas alternadas)
    body_font = Font(color="000000")
    body_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=len(encabezados)), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment
            cell.fill = alternating_colors[index % 2]  # Alterna colores de fondo

    # Ajustar márgenes y columnas
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 18
    worksheet.column_dimensions['I'].width = 15

    worksheet.sheet_view.showGridLines = False  # Oculta las líneas de cuadrícula

    # Ajustar los márgenes
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    workbook.save(response)
    return response



@login_required(login_url="login")
def entregados_folios(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        start_pedido = request.POST['start_pedido']
        end_pedido = request.POST['end_pedido']

        query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE,"
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PENUM >= '" + start_pedido + "' "
	            + "AND P.PENUM <= '" + end_pedido + "' "
                + "AND B.PEPAR0 >= '8' "
                + "AND B.PEPAR0 <= '8ZZZZZZ' "
                + "AND B.PENUM >= 'E' "
                + "AND B.PENUM <= 'E999999999' "
                + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) <= (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

        results = run_query(query) 

        resultado = list(results)

        return render(request, 'entregados.html',{
            'title':'Listado de Pedidos',
            'resultado':resultado,
            'leyenda':True,
            'fecha_Ini':start_pedido,
            'fecha_Fin':end_pedido,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
        return render(request, 'entregados.html',{
            'title':'Listado de Pedidos',
            'leyenda':False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required(login_url="login")
def entregados_fecha(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        start_fecha = request.POST['start_date']
        end_fecha = request.POST['end_date']

    if start_fecha and end_fecha:
        query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE,"
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PEVENCE >= '" + start_fecha + "' "
	            + "AND B.PEVENCE <= '" + end_fecha + "' "
                + "AND B.PEPAR0 >= '8' "
                + "AND B.PEPAR0 <= '8ZZZZZZ' "
                + "AND B.PENUM >= 'E' "
                + "AND B.PENUM <= 'E999999999' "
                + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) <= (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

        results = run_query(query) 

        resultado = list(results)

        return render(request, 'entregados.html',{
            'title':'Listado de Pedidos',
            'resultado':resultado,
            'leyenda':True,
            'fecha_Ini':start_fecha,
            'fecha_Fin':end_fecha,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
        return render(request, 'entregados.html',{
            'title':'Listado de Pedidos',
            'leyenda':False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })


@login_required(login_url="login")
def dashboard_Bord(request):
    date_now = datetime.datetime.now().date()
    last_year = date_now - datetime.timedelta(days=365)
    

    presente = date_now.strftime('%Y-%m-%d')
    pasado = last_year.strftime('%Y-%m-%d')

    fecha_actual = 0
    anterior = 0

    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query = (f"SELECT COUNT(*), (SELECT COUNT(*) FROM FPENC B WHERE B.PEFECHA = '" + pasado + "' AND B.PEPAR0 >= '7' "
             + "AND B.PEPAR0 <= '8ZZZZZZ' AND B.PENUM >= 'E' AND B.PENUM <= 'E999999999') FROM FPENC B "
             + "WHERE B.PEFECHA = '" + presente + "' AND B.PEPAR0 >= '7' AND B.PEPAR0 <= '8ZZZZZZ' AND B.PENUM >= 'E' AND B.PENUM <= 'E999999999'")

    results = run_query(query)
    resultado_dash = list(results)

    for result in resultado_dash:
        fecha_actual = result[0]
        anterior = result[1]

    borlazzar = 0
    fernava = 0
    francoalfaro = 0
    cruzacosta = 0
    graftext = 0
     
    query = (f"SELECT V.PRVNOM, SUM((SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ)- "
            + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ)) AS PIEZAS FROM FPENC P "
            + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
            + "WHERE B.PEPAR0 >= '7' AND B.PEPAR0 <= '8ZZZZZZ' AND B.PENUM >= 'E' AND B.PENUM <= 'E999999999' "
            + "AND B.PEINICIAL = 0 AND (SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) > "
            + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) GROUP BY V.PRVNOM;")
     
    results = run_query(query)
    results_das = list (results)

    for resultado in results_das:
        provider_name = resultado[0]
        if provider_name == "Bordadora Lazzar":
            borlazzar += 1
        elif provider_name == "Fernandez Navarro": 
            fernava += 1
        elif provider_name == "Franco Alfaro":
            francoalfaro += 1
        elif provider_name == "Cruz Acosta":
            cruzacosta += 1
        elif provider_name == "M. grafíca textil SA de CV":
            graftext += 1

    return render(request, 'dashboard_Bord.html', {
        'result': resultado_dash,
        'fecha_actual': fecha_actual,
        'anterior': anterior,
        'presente': presente,
        'date_now': date_now,
        'pasado': pasado,
        'last_year': last_year,

        'resultado': results_das,
        'borlazzar':borlazzar,
        'fernava':fernava,
        'francoalfaro':francoalfaro,
        'cruzacosta':cruzacosta,
        'graftext':graftext,

        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })


@login_required(login_url="login")
def buscar_ventas(request):
    usuario = CorreosLazzar.objects.get(correo=request.user.email)
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups
    if request.method == 'POST':
        search_pedido = request.POST.get('search_pedido', None)

        query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE,"
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC B "
                + "INNER JOIN FPENC P ON P.PENUM = B.PEOTROTXT "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PENUM = '" + search_pedido + "';")

        results = run_query(query) 

        resultado = list(results)

        return render(request, 'buscar_ventas.html',{
            'title':'INFORMACIÓN',
            'resultado':resultado,
            'leyenda':True,
            'search_pedido':search_pedido,
            'usuario': usuario,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
        return render(request, 'buscar_ventas.html',{
            'title':'INFORMACIÓN',
            'leyenda':False,
            'usuario': usuario,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

login_required(login_url="login")
def pedidos_dia (request):

    fecha_actual = datetime.datetime.now().date()
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    query = ( f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE, "
    + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
    + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
    + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
    + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
    + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
    + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
    + "FROM FPENC P "
    + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
    + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
    + "WHERE B.PEPAR0 >= '8' "
    + "AND B.PEPAR0 <= '8ZZZZZZ' "
    + "AND B.PENUM >= 'E' "
    + "AND B.PENUM <= 'E999999999' "
    + "AND B.PEINICIAL = 0 "
    + f"AND B.PEFECHA = '{fecha_actual}' "  # Agregar la condición de fecha actual
    + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L "
    + "WHERE L.PESEQ = B.PESEQ) > (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

    results = run_query(query) 

    resultado = list(results)

    return render(request, 'pedidos_dia.html',{
        'resultado':resultado,
        'leyenda':True,
        'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
        'acceso_total': acceso_total,
        'mesa_control': mesa_control, # End Admin-Groups
    })

@login_required(login_url="login")
def productividad(request):
    ventas = request.user.groups.filter(name='ventas').exists() #Admin-Groups Jesus Ibarra 19/12/2023
    acceso_total = request.user.groups.filter(name='acceso_total').exists()
    mesa_control = request.user.groups.filter(name='mesa_control').exists() # End Admin-Groups

    if request.method == 'POST':
        start_prod = request.POST['start_prod']

        query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE,"
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PEDESDE = '" + start_prod + "' "
                + "AND B.PEPAR0 >= '8' "
                + "AND B.PEPAR0 <= '8ZZZZZZ' "
                + "AND B.PENUM >= 'E' "
                + "AND B.PENUM <= 'E999999999' "
                + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) <= (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

        results = run_query(query) 

        resultado = list(results)

        return render(request, 'historial.html',{
            'resultado':resultado,
            'leyenda':True,
            'fecha_Ini':start_prod,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })
    
    else:
        return render(request, 'historial.html',{
            'leyenda':False,
            'ventas': ventas, #Admin-Groups Jesus Ibarra 19/12/2023
            'acceso_total': acceso_total,
            'mesa_control': mesa_control, # End Admin-Groups
        })

@login_required(login_url="login") 
def descargar_excel_productividad(request):
    
    query = (f"SELECT P.PENUM, P.PEFECHA, B.PENUM, V.PRVNOM, B.PEFECHA, B.PEDESDE, B.PEVENCE,"
                + "(SELECT C.COML1 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_1,"
                + "(SELECT C.COML2 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_2,"
                + "(SELECT C.COML3 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_3,"
                + "(SELECT C.COML4 FROM FCOMENT C WHERE C.COMSEQFACT = concat(10,B.PESEQ)) AS COMENT_4,"
                + "(SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS PIEZAS,"
                + "(SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) AS SURTIDO "
                + "FROM FPENC P "
                + "INNER JOIN FPENC B ON B.PEOTROTXT = P.PENUM "
                + "INNER JOIN FPRV V ON V.PRVSEQ = B.PRVSEQ "
                + "WHERE B.PEDESDE = DATE_SUB(CURRENT_DATE, INTERVAL 2 DAY) "
                + "AND B.PEPAR0 >= '8' "
                + "AND B.PEPAR0 <= '8ZZZZZZ' "
                + "AND B.PENUM >= 'E' "
                + "AND B.PENUM <= 'E999999999' "
                + "AND (SELECT SUM(L.PLCANT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ) <= (SELECT SUM(L.PLSURT) FROM FPLIN L WHERE L.PESEQ = B.PESEQ);")

    results = run_query(query)
    resultado = list(results)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=pedidos_pendientes.xlsx'

    workbook = Workbook()
    worksheet = workbook.active

    encabezados = ["PEDIDO", "FECHA", "ORDEN", "PROVEEDOR", "FECHA DE SURTIDO", "RECEPCIÓN DE LA ORDEN", "FECHA DE ENTREGA", "SURTIDO", "ESTATUS", "RECIBIÓ", "DESEBRO", "PIEZAS", "ENTREGADO"]
    worksheet.append(encabezados)

    # Aplicar formato a los encabezados
    header_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    header_font = Font(color="000000", bold=True)
    header_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = header_alignment

    alternating_colors = [PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
                          PatternFill(start_color="EFEFEF", end_color="EFEFEF", fill_type="solid")]

    for fila in resultado:
        cleaned_row = ["" if (cell == "1900/12/31" or cell == "31/12/1900") else cell for cell in fila]
        worksheet.append(cleaned_row)

    # Aplicar estilos al cuerpo de la tabla (filas alternadas)
    body_font = Font(color="000000")
    body_border = Border(top=Side(style="thin"), bottom=Side(style="thin"), left=Side(style="thin"), right=Side(style="thin"))
    body_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)  # Texto centrado

    for index, row in enumerate(worksheet.iter_rows(min_row=2, min_col=1, max_col=len(encabezados)), start=2):
        for cell in row:
            cell.font = body_font
            cell.border = body_border
            cell.alignment = body_alignment
            cell.fill = alternating_colors[index % 2]  # Alterna colores de fondo

    # Ajustar márgenes y columnas
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['D'].width = 20
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 12
    worksheet.column_dimensions['G'].width = 12
    worksheet.column_dimensions['H'].width = 18
    worksheet.column_dimensions['I'].width = 15

    worksheet.sheet_view.showGridLines = False  # Oculta las líneas de cuadrícula

    # Ajustar los márgenes
    worksheet.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    workbook.save(response)
    return response